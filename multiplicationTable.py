#!python

import os, openpyxl, sys
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter, column_index_from_string

wb=openpyxl.Workbook()
sheet = wb.active
sheet.title = 'Multiplication Table'

fontObj = Font(bold=True)

def multiplication(num, filename):
	#populate the A column
	for i in range(1, int(num)+1): 
		sheet['A' + str(i+1)] = i
		sheet['A' + str(i+1)].font = fontObj

	#populate the first row	
	for r in range(1,int(num)+1): 
		col = get_column_letter(r+1)
		columns = col+str(1)
		sheet[columns].value = r
		sheet[columns].font = fontObj

	#populate the cells with the products
	for k in range(1, int(num)+1):
		for j in range(1, int(num)+1):
			sheet.cell(row=k+1, column=j+1).value = k*j
	print('Multiplication Table is ready to be viewed')


		
	wb.save(filename + '.xlsx')
print(sys.argv[1])
userNumber = sys.argv[1]
userFileName = sys.argv[2]

if len(sys.argv) < 2:
	print('Please enter a number, followed by a word that will be used as a filename')	
else:
	multiplication(userNumber, userFileName) 


